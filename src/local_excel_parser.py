from __future__ import annotations

import math
import re
import shutil
import subprocess
import tempfile
import hashlib
from dataclasses import dataclass
from datetime import datetime
from html import escape
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}


@dataclass
class ExcelSheet:
    name: str
    content_type: str
    start_line: int
    end_line: int
    content: str
    rows: list[list[str]]
    char_count: int
    table_char_count: int
    preview_rows: list[list[str]]


@dataclass
class ExcelSearchHit:
    line_no: int
    text: str
    keyword: str


@dataclass
class ExcelCellMatch:
    sheet: str
    row: int
    column: int
    cell: str
    header: str
    row_label: str
    value: str


class LocalExcelParseService:
    def __init__(self, path: str | Path, file_id: str | None = None) -> None:
        self.path = Path(path).expanduser()
        self._parse_path = _resolve_excel_parse_path(self.path)
        self.file_id = file_id or self.path.stem
        self.file_name = self.path.name
        self.lines: list[str] = []
        self.sheets: list[ExcelSheet] = []
        self._parse()

    def extract_file_summary_text(self) -> str:
        headings = "\n".join(
            f'<sheet name="{sheet.name}" contentType="{sheet.content_type}">'
            for sheet in self.sheets
        )
        return self._wrap_result(
            "# 总体摘要\n\n"
            f"该文档共{len(self.lines)}行,{self._total_chars()}字符,整体分为正文部分。\n\n"
            "# 正文部分:\n"
            "## 摘要信息\n"
            f"字符总数:{self._total_chars()},总行数:{len(self.lines)},"
            f"行索引范围:1~{len(self.lines)}\n"
            "## 章节标题信息\n"
            f"{headings}"
        )

    def extract_chapter_summary(self, chapters: str) -> str:
        parts: list[str] = []
        missing: list[str] = []
        for query in self._split_csv(chapters):
            sheet = self._find_sheet(query)
            if sheet is None:
                missing.append(query)
                continue
            parts.append(self._sheet_summary(sheet))
        if not parts:
            return self._wrap_tip(f"未找到{','.join(missing)}\n对应的章节，请核实编号或名称")
        if missing:
            return self._wrap_tip_result(
                f"未找到{','.join(missing)}\n对应的章节，请核实编号或名称",
                "\n".join(parts),
            )
        return self._wrap_result("\n".join(parts))

    def extract_chapter_content(self, chapters: str) -> str:
        parts: list[str] = []
        missing: list[str] = []
        for query in self._split_csv(chapters):
            sheet = self._find_sheet(query)
            if sheet is None:
                missing.append(query)
                continue
            parts.extend(self.lines[sheet.start_line - 1 : sheet.end_line])
        if not parts:
            return self._wrap_tip(f"未找到{','.join(missing)}\n对应的章节，请核实编号或名称")
        return self._wrap_result("\n".join(parts))

    def read_lines(self, start_line: int, end_line: int) -> str:
        if not self.lines:
            return self._wrap_result("")
        start = max(1, start_line)
        end = min(len(self.lines), end_line)
        if start > end:
            return self._wrap_result("")
        return self._wrap_result("\n".join(self.lines[start - 1 : end]))

    def search_content(
        self,
        keywords: str,
        page_no: int | None = None,
        page_size: int | None = None,
    ) -> str:
        keys = self._split_csv(keywords)
        hits: list[ExcelSearchHit] = []
        for line_no, line in enumerate(self.lines, start=1):
            for key in keys:
                if key and key in line:
                    hits.append(ExcelSearchHit(line_no=line_no, text=line, keyword=key))
                    break
        page_no = max(1, page_no or 1)
        page_size = max(1, page_size or 20)
        total_pages = math.ceil(len(hits) / page_size) if hits else 0
        start = (page_no - 1) * page_size
        selected = hits[start : start + page_size]
        body = "\n".join(f"{hit.line_no} {hit.text}" for hit in selected)
        tip = (
            f"搜索结果总条数:{len(hits)}\n"
            f"当前页码:{page_no}\n"
            f"每页条数:{page_size}\n"
            f"总页数:{total_pages}"
        )
        return self._wrap_tip_result(tip, body)

    def extract_table_list(self, chapters: str) -> str:
        body = "\n".join(
            f'<sheet name="{sheet.name}" contentType="{sheet.content_type}" rows="{len(sheet.rows)}" cols="{max((len(row) for row in sheet.rows), default=0)}">'
            for sheet in self.sheets
        )
        return self._wrap_result(body or '""')

    def extract_table_content(
        self,
        table_title: str,
        row_scope: str | None = None,
        row_keywords: str | None = None,
        col_keywords: str | None = None,
    ) -> str:
        sheets = self._matching_sheets(table_title)
        if not sheets:
            return self._wrap_tip("未找到对应工作表，请核实 tableTitle 是否准确")
        matches: list[ExcelCellMatch] = []
        for sheet in sheets:
            matches.extend(
                _match_sheet_cells(
                    sheet,
                    row_keywords=row_keywords or "",
                    col_keywords=col_keywords or "",
                    row_scope=row_scope,
                )
            )
        if not matches:
            return self._wrap_tip("未检索到相关内容，请核实行关键词、列关键词或筛选范围是否准确")
        body = "\n".join(_format_cell_match(match) for match in matches[:80])
        tip = (
            f"检索结果总条数:{len(matches)}\n"
            f"返回条数:{min(len(matches), 80)}\n"
            "结果格式: 工作表!单元格 行标签 / 列标题 = 单元格值"
        )
        return self._wrap_tip_result(tip, body)

    def _parse(self) -> None:
        workbook = load_workbook(str(self._parse_path), data_only=True, read_only=False)
        visible_worksheets = [
            worksheet for worksheet in workbook.worksheets if worksheet.sheet_state == "visible"
        ]
        for index, worksheet in enumerate(visible_worksheets):
            rows = _sheet_rows(worksheet)
            if not rows:
                continue
            content_type = _sheet_content_type(worksheet, rows)
            marker = f'<sheet name="{worksheet.title}" contentType="{content_type}">'
            start_line = self._append_line(marker)
            content = (
                _sheet_to_markdown(rows)
                if content_type == "markdown"
                else _sheet_to_html(worksheet, rows)
            )
            for line in content.splitlines():
                self._append_line(line)
            end_line = len(self.lines)
            self.sheets.append(
                ExcelSheet(
                    name=worksheet.title,
                    content_type=content_type,
                    start_line=start_line,
                    end_line=end_line,
                    content=content,
                    rows=rows,
                    char_count=len("\n".join(self.lines[start_line - 1 : end_line])),
                    table_char_count=len(content),
                    preview_rows=[row[:2] for row in rows[:50]],
                )
            )
            if index < len(visible_worksheets) - 1:
                self._append_line("")

    def _sheet_summary(self, sheet: ExcelSheet) -> str:
        preview = "\n".join(
            f"{idx}: {' | '.join(row)}" for idx, row in enumerate(sheet.preview_rows, start=1)
        )
        return (
            f'<sheet name="{sheet.name}" contentType="{sheet.content_type}">\n'
            f"\t字符总数:{sheet.char_count},其中表格字符总数:{sheet.table_char_count}\n"
            f"\t总行数:{max(0, sheet.end_line - sheet.start_line + 1)},"
            f"行索引范围:{sheet.start_line}~{sheet.end_line}\n"
            "\t# 表格摘要:\n"
            f"\t{sheet.content}\n\n"
            "# 前两列预览(第1、2列):\n"
            f"{preview}"
        )

    def _find_sheet(self, query: str) -> ExcelSheet | None:
        normalized = _normalize(query)
        for sheet in self.sheets:
            candidates = {
                sheet.name,
                str(self.sheets.index(sheet) + 1),
                f'<sheet name="{sheet.name}" contentType="{sheet.content_type}">',
            }
            if normalized in {_normalize(item) for item in candidates}:
                return sheet
        for sheet in self.sheets:
            if normalized and normalized in _normalize(sheet.name):
                return sheet
        return None

    def _matching_sheets(self, query: str) -> list[ExcelSheet]:
        if not query.strip():
            return self.sheets
        exact = self._find_sheet(query)
        if exact is not None:
            return [exact]
        normalized = _normalize(query)
        return [
            sheet
            for sheet in self.sheets
            if normalized and (normalized in _normalize(sheet.content) or normalized in _normalize(sheet.name))
        ]

    def _append_line(self, text: str) -> int:
        self.lines.append(text)
        return len(self.lines)

    def _wrap_result(self, text: str) -> str:
        return f"{self._meta()}\n<result>\n{text}</result>"

    def _wrap_tip(self, text: str) -> str:
        return f"{self._meta()}\n<tip>{text}</tip>"

    def _wrap_tip_result(self, tip: str, result: str) -> str:
        return f"{self._meta()}\n<tip>\n{tip}\n</tip>\n<result>\n{result}</result>"

    def _meta(self) -> str:
        return (
            "<meta>\n"
            f"<file-id>{self.file_id}</file-id>\n"
            f"<file-name>{self.file_name}</file-name>\n"
            "</meta>"
        )

    def _total_chars(self) -> int:
        return len("\n".join(self.lines))

    def _split_csv(self, value: str) -> list[str]:
        return [part.strip() for part in value.split(",") if part.strip()]


def _resolve_excel_parse_path(path: Path) -> Path:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return path
    if suffix == ".xls":
        converted = _convert_xls_to_xlsx(path)
        if converted is not None:
            return converted
        raise ValueError(f"无法自动转换 .xls 文件: {path}. 请安装 LibreOffice 后重试。")
    return path


def _convert_xls_to_xlsx(path: Path) -> Path | None:
    out_dir = _xls_conversion_cache_dir(path)
    out_dir.mkdir(parents=True, exist_ok=True)
    cached = out_dir / f"{path.stem}.xlsx"
    if cached.exists():
        return cached
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        return None
    subprocess.run(
        [
            soffice,
            "-env:UserInstallation=file://" + str(out_dir / "profile"),
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out_dir),
            str(path),
        ],
        check=False,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    candidate = out_dir / f"{path.stem}.xlsx"
    return candidate if candidate.exists() else None


def _xls_conversion_cache_dir(path: Path) -> Path:
    stat = path.stat()
    digest = hashlib.sha256(
        f"{path.resolve()}:{stat.st_mtime_ns}:{stat.st_size}".encode("utf-8")
    ).hexdigest()[:16]
    return Path(tempfile.gettempdir()) / "document_parser_xls_cache" / digest


def _sheet_rows(worksheet: Worksheet) -> list[list[str]]:
    max_row, max_col = _used_range(worksheet)
    if max_row == 0 or max_col == 0:
        return []
    rows: list[list[str]] = []
    header_cells = [
        _cell_text(cell)
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col))
    ]
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        rows.append(
            [
                _cell_text(cell, header=header_cells[index] if index < len(header_cells) else "")
                for index, cell in enumerate(row)
            ]
        )
    return _trim_empty_edges(rows)


def _used_range(worksheet: Worksheet) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if _cell_text(cell):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def _trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(cell for cell in rows[-1]):
        rows.pop()
    if not rows:
        return []
    width = max(len(row) for row in rows)
    while width > 0 and not any(len(row) >= width and row[width - 1] for row in rows):
        width -= 1
    return [row[:width] + [""] * max(0, width - len(row)) for row in rows]


def _cell_text(cell: Cell, header: str = "") -> str:
    value = cell.value
    if value is None:
        return ""
    normalized_header = _normalize(header)
    if _looks_like_date_header(normalized_header):
        date_text = _date_cell_text(value)
        if date_text:
            return date_text
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _looks_like_date_header(normalized_header: str) -> bool:
    return normalized_header in {"日期", "时间", "月份", "月度"} or normalized_header.endswith("日期")


def _date_cell_text(value: object) -> str:
    date_value: datetime | None = None
    if isinstance(value, datetime):
        date_value = value
    elif isinstance(value, int | float):
        if 20000 <= float(value) <= 90000:
            try:
                converted = from_excel(value)
            except Exception:
                converted = None
            if isinstance(converted, datetime):
                date_value = converted
    if date_value is None:
        return ""
    date = date_value.date()
    return f"{date.isoformat()}（{date.month}月）"


def _sheet_content_type(worksheet: Worksheet, rows: list[list[str]]) -> str:
    if worksheet.merged_cells.ranges:
        return "html"
    if len(rows) > 12 and max((len(row) for row in rows), default=0) <= 4:
        return "markdown"
    return "html"


def _sheet_to_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = "|" + "|".join(normalized[0]) + "|"
    body = ["|" + "|".join(row) + "|" for row in normalized[1:]]
    return "\n".join([header, *body])


def _sheet_to_html(worksheet: Worksheet, rows: list[list[str]]) -> str:
    merge_map = _merged_cell_map(worksheet)
    headers = rows[0] if rows else []
    html_rows: list[str] = []
    for row_idx, row in enumerate(rows, start=1):
        cells: list[str] = []
        for col_idx, value in enumerate(row, start=1):
            merged = merge_map.get((row_idx, col_idx))
            if merged == "skip":
                continue
            attrs: list[str] = [
                f'cell="{get_column_letter(col_idx)}{row_idx}"',
            ]
            if row_idx > 1 and col_idx <= len(headers) and headers[col_idx - 1]:
                attrs.append(f'header="{_format_html_attr(headers[col_idx - 1])}"')
            label = _row_label(row, col_idx)
            if label:
                attrs.append(f'label="{_format_html_attr(label)}"')
            if isinstance(merged, tuple):
                rowspan, colspan = merged
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
            attr_text = " " + " ".join(attrs) if attrs else ""
            cells.append(f"<td{attr_text}>{_format_html_cell(value)}</td>")
        html_rows.append(
            f'<tr sheet="{_format_html_attr(worksheet.title)}" row="{row_idx}">'
            f"{''.join(cells)}</tr>"
        )
    return "<table>\n" + "\n".join(html_rows) + "\n</table>"


def _match_sheet_cells(
    sheet: ExcelSheet,
    row_keywords: str,
    col_keywords: str,
    row_scope: str | None,
) -> list[ExcelCellMatch]:
    if not sheet.rows:
        return []
    row_terms = _split_keywords(row_keywords)
    col_terms = _split_keywords(col_keywords)
    row_indexes = _row_indexes(sheet.rows, row_terms, row_scope)
    col_indexes = _col_indexes(sheet.rows, col_terms)
    if row_terms and not col_terms and _row_terms_match_headers(sheet.rows, row_terms):
        col_indexes = _col_indexes(sheet.rows, row_terms)
        row_indexes = list(range(1, len(sheet.rows)))
    if not row_indexes:
        row_indexes = list(range(1, len(sheet.rows))) if col_indexes else []
    if not col_indexes:
        col_indexes = list(range(max((len(row) for row in sheet.rows), default=0)))
    matches: list[ExcelCellMatch] = []
    headers = sheet.rows[0]
    for row_idx in row_indexes:
        if row_idx <= 0 or row_idx >= len(sheet.rows):
            continue
        row = sheet.rows[row_idx]
        row_label = row[0] if row else ""
        for col_idx in col_indexes:
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if not value:
                continue
            header = headers[col_idx] if col_idx < len(headers) else ""
            matches.append(
                ExcelCellMatch(
                    sheet=sheet.name,
                    row=row_idx + 1,
                    column=col_idx + 1,
                    cell=f"{get_column_letter(col_idx + 1)}{row_idx + 1}",
                    header=header,
                    row_label=row_label,
                    value=value,
                )
            )
    return matches


def _row_indexes(
    rows: list[list[str]],
    row_terms: list[str],
    row_scope: str | None,
) -> list[int]:
    scoped = _parse_row_scope(row_scope, len(rows))
    if not row_terms:
        return scoped
    indexes = [
        idx
        for idx in scoped
        if idx > 0 and _all_terms_match(" ".join(rows[idx]), row_terms)
    ]
    if indexes:
        return indexes
    month_terms = [_month_term(term) for term in row_terms]
    month_terms = [term for term in month_terms if term]
    if not month_terms:
        return []
    return [
        idx
        for idx in scoped
        if idx > 0 and any(month in _normalize(" ".join(rows[idx])) for month in month_terms)
    ]


def _col_indexes(rows: list[list[str]], col_terms: list[str]) -> list[int]:
    if not rows:
        return []
    headers = rows[0]
    if not col_terms:
        return []
    indexes = [
        idx
        for idx, header in enumerate(headers)
        if header and _all_terms_match(header, col_terms)
    ]
    if indexes:
        return indexes
    return [
        idx
        for idx, header in enumerate(headers)
        if header and _any_term_match(header, col_terms)
    ]


def _row_terms_match_headers(rows: list[list[str]], row_terms: list[str]) -> bool:
    if not rows:
        return False
    return any(_all_terms_match(header, row_terms) for header in rows[0])


def _parse_row_scope(row_scope: str | None, row_count: int) -> list[int]:
    if not row_scope:
        return list(range(row_count))
    indexes: list[int] = []
    for part in re.split(r"[,，]", row_scope):
        part = part.strip()
        if not part:
            continue
        match = re.match(r"^(\d+)\s*[-~]\s*(\d+)$", part)
        if match:
            start, end = int(match.group(1)), int(match.group(2))
            indexes.extend(range(max(1, start) - 1, min(row_count, end)))
            continue
        if part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < row_count:
                indexes.append(idx)
    return sorted(set(indexes))


def _split_keywords(value: str) -> list[str]:
    raw_terms = [part.strip() for part in re.split(r"[,，\\s]+", value) if part.strip()]
    terms: list[str] = []
    for term in raw_terms:
        terms.append(term)
        month = _month_term(term)
        if month:
            terms.append(month)
    seen: set[str] = set()
    deduped: list[str] = []
    for term in terms:
        normalized = _normalize(term)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(term)
    return deduped


def _all_terms_match(text: str, terms: list[str]) -> bool:
    normalized = _normalize(text)
    return all(_normalize(term) in normalized for term in terms if _normalize(term))


def _any_term_match(text: str, terms: list[str]) -> bool:
    normalized = _normalize(text)
    return any(_normalize(term) in normalized for term in terms if _normalize(term))


def _month_term(value: str) -> str:
    match = re.search(r"(\d{1,2})\s*月", value)
    if match:
        return f"{int(match.group(1))}月"
    return ""


def _format_cell_match(match: ExcelCellMatch) -> str:
    header = re.sub(r"\s+", " ", match.header).strip()
    row_label = re.sub(r"\s+", " ", match.row_label).strip()
    return (
        f"{match.sheet}!{match.cell} "
        f'row="{row_label}" header="{header}" = {match.value}'
    )


def _row_label(row: list[str], col_idx: int) -> str:
    if col_idx <= 1:
        return ""
    first_value = row[0] if row else ""
    if _looks_like_date_text(first_value):
        return first_value
    for value in reversed(row[: col_idx - 1]):
        normalized = _normalize(value)
        if not normalized:
            continue
        if normalized in {"/", "-", "--", "—"}:
            continue
        if re.fullmatch(r"\d+(\.\d+)?", normalized):
            continue
        return value
    return ""


def _looks_like_date_text(value: str) -> bool:
    return bool(re.search(r"\d{4}-\d{2}-\d{2}", value))


def _merged_cell_map(worksheet: Worksheet) -> dict[tuple[int, int], tuple[int, int] | str]:
    mapping: dict[tuple[int, int], tuple[int, int] | str] = {}
    for merged in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        mapping[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                mapping[(row, col)] = "skip"
    return mapping


def _format_html_cell(value: str) -> str:
    return escape(value).replace("\n", "<br>")


def _format_html_attr(value: str) -> str:
    return escape(value, quote=True).replace("\n", " ")


def _normalize(value: str) -> str:
    return re.sub(r"\s+", "", value).strip()
